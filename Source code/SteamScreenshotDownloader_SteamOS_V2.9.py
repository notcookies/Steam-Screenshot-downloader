import os
import re
import sys
import time
import random
import hashlib
import openpyxl
#import win32con
import requests
import threading
#import win32file
#import pythoncom
#import pywintypes
import pandas as pd
#import win32timezone
import tkinter as tk
from PIL import Image
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from urllib.parse import unquote
from openpyxl import load_workbook
from tkinter import ttk, filedialog
from urllib.parse import urlparse, urlunparse
#from selenium.webdriver.chrome.options import Options
#from selenium.webdriver.chrome.service import Service
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
os.environ["MOZ_DISABLE_CONTENT_SANDBOX"] = "1"


# Constants
# CHROME_PATH = os.path.join(os.getcwd(), "chrome", "chrome.exe")
# CHROMEDRIVER_PATH = os.path.join(os.getcwd(), "chrome", "chromedriver.exe")
# CHROME_PATH = "/home/jinx/chrome/chrome"
# CHROMEDRIVER_PATH = "/home/jinx/chrome/chromedriver"
# PROFILE_DIR = "Default"
# Firefox_profile_path = "/home/jinx/.mozilla/firefox/wpljomkx.default-release"
# Firefox_path         = "/home/jinx/firefoox/firefox"
#No Content-Disposition
CONTENT_TYPE_TO_EXT = {
    'image/jpeg': '.jpg',
    'image/jpg': '.jpg',
    'image/png': '.png',
    'image/gif': '.gif',
    'image/webp': '.webp',
}
# Redirect stdout to Text widget
class TextRedirector:
    def __init__(self, widget):
        self.widget = widget

    def write(self, message):
        self.widget.after(0, self.widget.insert, tk.END, message)
        self.widget.after(0, self.widget.see, tk.END)

    def flush(self):
        pass

#Chrome init
# def init_chrome(user_data_path=None):
#     options = Options()
#     # options.binary_location = CHROME_PATH
#     # options.add_argument(f"--user-data-dir={user_data_path}")
#     # options.add_argument(f"--profile-directory={PROFILE_DIR}")
#     # options.add_argument("--log-level=3")
#     # service = Service(CHROMEDRIVER_PATH)
#     options.profile = Firefox_profile_path 
#     options.binary_location = Firefox_path
#     #return webdriver.Chrome(service=service, options=options)
#     return webdriver.Firefox(options=options)

# Detect Firefox default profile folder
def detect_firefox_default_profile_folder():
    base_path = os.path.expanduser("~/.mozilla/firefox")
    if not os.path.exists(base_path):
        return ""

    for folder in os.listdir(base_path):
        full_path = os.path.join(base_path, folder)
        if os.path.isdir(full_path) and re.match(r".*\.default-release$", folder):
            return full_path

    return ""

def init_chrome(user_data_path=None):
    options = Options()
    if not user_data_path:
        user_data_path = detect_firefox_default_profile_folder()
        if not user_data_path:
            return None
    cwd = os.getcwd()
    firefox_path =  "/usr/lib/firefox/firefox"
    driver_path  = os.path.join(cwd, "geckodriver")
    options.add_argument("--allow-downgrade")
    options.add_argument(f"-profile")
    options.add_argument(f"{user_data_path}")
    options.binary_location = firefox_path
    service = Service(executable_path=driver_path)

    return webdriver.Firefox(service = service, options=options)


# Get screenshot links
def get_screenshot_links(driver, steam_id, page, appid=None, screenshot=True, favorite=False, custom_url=True):
    if custom_url:
        url_base = "https://steamcommunity.com/id/"
    else:
        url_base = "https://steamcommunity.com/profiles/"

    if favorite:
        browsefilter = "&browsefilter=myfavorites"
    else:
        browsefilter = "&browsefilter=myfiles"

    if screenshot:
        base_url = f"{url_base}{steam_id}/screenshots/?p={page}&sort=oldestfirst{browsefilter}&view=grid&privacy=30"
    else:
        base_url = f"{url_base}{steam_id}/images/?p={page}&sort=oldestfirst{browsefilter}&view=grid&privacy=30"

    if appid:
        base_url += f"&appid={appid}"

    for _ in range(6):
        time.sleep(random.uniform(1, 2))
        driver.get(base_url)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        links = [a['href'] for a in soup.select('a[href*="//steamcommunity.com/sharedfiles/filedetails/"]')]
        if links:
            print(f"[Page {page}] Found {len(links)} screenshot links.")
            print(f"Fetching the original resolution URLs of the {len(links)} screenshots, please wait ...")
            indexed_links = [(i, link) for i, link in enumerate(links)]
            return indexed_links
    print(f"[Page {page}] No screenshot links found.")
    return []


# Chrome.get_cookies
def extract_steam_cookies_from_driver(driver, retries=3, delay=2):
    for attempt in range(1, retries + 1):
        cookies = driver.get_cookies()
        cookie_dict = {c['name']: c['value'] for c in cookies}
        # check cookie
        if all(k in cookie_dict for k in ["steamLoginSecure", "sessionid"]):
            print("✅ Steam cookies extracted successfully.")
            return cookie_dict
        time.sleep(delay)
    print("❌ Failed to extract required cookies after 3 attempts.")
    return {}

# Get image URL
# change driver.get to request_html
def get_img_url_from_html(link, cookies):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0"
        }
        r = requests.get(link, headers=headers, cookies=cookies, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        tag = soup.select_one('a[href*="//images.steamusercontent.com/ugc/"]')
        if tag:
            link = urlunparse(urlparse(tag['href'])._replace(query=''))
            link_full = urlunparse(urlparse(tag['href']))
            return link, link_full
    except Exception as e:
        print(f"Error fetching {link}: {e}")
    return None

# fetch requests img_urls concurrently with original index
def fetch_img_urls_concurrently_requests(indexed_links, cookies, page, processes, max_retries=3):
    total_links = len(indexed_links)

    for attempt in range(max_retries):

        def worker(index_link_tuple):
            index, link = index_link_tuple
            result = get_img_url_from_html(link, cookies)
            if result:
                image_url, image_url_query = result
                return (index, link, image_url, image_url_query)
            return None
        
        results = []
        with ThreadPoolExecutor(max_workers=processes) as executor:
            futures = {executor.submit(worker, indexed_link): indexed_link for indexed_link in indexed_links}
            for future in as_completed(futures):
                result = future.result(timeout=15)
                if result:
                    results.append(result)

        success_count = len(results)
        print(f"[Page {page}] Got {success_count}/{total_links} image links.")
        
        if success_count == total_links:

            # current sort by original index
            # results.sort(key=lambda x: x[0])  
            # img_links = [(link, img_url, image_url_query) for index, link, img_url, image_url_query in results]           
            print(f"[Page {page}] ✅ Retrieved {len(results)} original image links.\n")
            return results
        else:
            missing = total_links - success_count
            print(f"[Page {page}] Missing {missing} links, retrying...")
            if attempt < max_retries - 1:
                time.sleep(2)
            else:
                print(f"[Page {page}] ❌ Final attempt failed, using {success_count} links.")
                # results.sort(key=lambda x: x[0]) 
                # img_links = [(link, img_url, image_url_query) for index, link, img_url, image_url_query in results]           
                print(f"[Page {page}] ✅ Retrieved {len(results)} original image links.\n")
                return results
            

def gen_img_id(url: str, length: int = 10) -> str:
    md5 = hashlib.md5(url.encode('utf-8')).hexdigest()
    return md5[:length]

def get_appid_filename_from_cd(debug, page_url, link, cd_header: str) -> tuple[str, str]:

    match = re.search(r'=\s*"?([^";]+)"?', cd_header)
    if debug:
        print(f"debug_Content-Disposition:{match}\n{link}\n{page_url}")
    if not match:
        print("Error field match filename Content-Disposition!")
        raise ValueError("No filename* found in Content-Disposition!")

    full_name = match.group(1)
    if debug:
        print(f"debug_match_Content-Disposition:{full_name} from{match}")
    idx = full_name.find("_screenshots_")
    # Handle Artwork filenames
    if idx == -1:
        appid = "Artwork"
        # # Extract filename
        # idx = full_name.find("_")
        # i = idx - 1
        # fname_chars = []
        # while i >= 0 and (full_name[i].isdigit() or full_name[i] == '_'):
        #     fname_chars.append(full_name[i])
        #     i -= 1
        # filename = ''.join(reversed(fname_chars))
        # # Append file extension
        # ext_match = re.search(r'\.(jpg|jpeg|png|gif|webp)', full_name, re.I)
        # if ext_match:
        #     filename += ext_match.group(0)
        # else:
        #     filename += ".jpg"
        if full_name.lower().startswith("utf-8''"):
                full_name = full_name[7:]
        filename = unquote(full_name)

        if debug:
            print(f"Artwork:\ndebug_match_Artwork_filename:{filename}\ndebug_match_Artwork_ext:{ext_match}\ndebug_match_Artwork_info:\nAppid:{appid}  Filename:{filename}")
        return appid, filename
    
    # Handle Screenshot filenames
    # Extract appid
    i = idx - 1
    app_chars = []
    while i >= 0 and (full_name[i].isdigit() or full_name[i] == '_'):
        app_chars.append(full_name[i])
        i -= 1
    appid = ''.join(reversed(app_chars))
    
    if appid == "0000":
        appid = "Error"

    # Extract filename
    j = idx + len("_screenshots_")
    fname_chars = []
    while j < len(full_name) and (full_name[j].isdigit() or full_name[j] in '_'or full_name[j] in '-' or full_name[j].isalnum()):
        fname_chars.append(full_name[j])
        j += 1
    filename = ''.join(fname_chars)

    # Append file extension
    ext_match = re.search(r'\.(jpg|jpeg|png|gif|webp)', full_name, re.I)
    if ext_match:
        filename += ext_match.group(0)
    else:
        filename += ".jpg"
    if debug:
            print(f"Screenshots:\ndebug_match_filename:{filename}\ndebug_match_ext:{ext_match}\ndebug_match_info:\nAppid:{appid}  Filename:{filename}")
    return appid, filename

# Download image
def download_img(page, link, img_url, image_url_query, save_dir, stop_flag, debug, idx):

    img_link_success = True
    same_cd = False  
    same_fname = False  
    file_written = False  
    missing_cd = False

    if stop_flag:
        img_link_success = False
        return {
            'Page': page, 'Index': idx if idx is not None else '?', 'Screenshot Link': link, 'Image Link': img_url,'Img Link Query': image_url_query,
            'Content Disposition': '', 'Content Type': '', 'File Name': '', 'App ID': '',
            'Missing Content Disposition': False, 'Same Content Disposition': False, 'Same File Name': False, 'File Written': False, 'Image Link Success': False
        }

    for attempt in range(3):
        try:
            missing_cd = False

            time.sleep(random.uniform(0.2, 0.5))
            r = requests.get(img_url,stream=True, timeout=5)
            r.raise_for_status()

            cd_header = r.headers.get('Content-Disposition', '')
            ct_header = r.headers.get('Content-Type', '')

            if not cd_header:

                missing_cd = True

                r = requests.get(image_url_query, stream=True, timeout=5)
                r.raise_for_status()
                ct_header = r.headers.get('Content-Type', '')

                ext = CONTENT_TYPE_TO_EXT.get(ct_header.lower().split(";")[0].strip(), ".jpg")

                unique_id = gen_img_id(img_url)

                cd_header = f"= 0000_screenshots_2000-01-01_{unique_id}{ext}"

                print(f"No Content-Disposition header found!\nError link:\n{img_url}\nFull url:\n{link}\nRenaming it {cd_header}")

            appid, new_fname = get_appid_filename_from_cd(debug, link, img_url,cd_header)

            folder = os.path.join(save_dir, appid, "screenshots")
            os.makedirs(folder, exist_ok=True)
            path = os.path.join(folder, new_fname)

            with open(path, "wb") as f:
                for chunk in r.iter_content(1024):
                    f.write(chunk)

            # confirm file is written
            file_written = os.path.exists(path) and os.path.getsize(path) > 0

            return {
                'Page': page,
                'Index': idx if idx is not None else '?',
                'Screenshot Link': link,
                'Image Link': img_url,
                'Img Link Query': image_url_query,
                'Content Disposition': cd_header,
                'Content Type': ct_header,
                'File Name': new_fname,
                'App ID': appid,
                'Missing Content Disposition': missing_cd,

                #check in run_downloader
                'Same Content Disposition': same_cd,
                'Same File Name': same_fname,

                'File Written': file_written,
                'Image Link Success': img_link_success
            }
        except Exception as e:
            if attempt == 2:
                img_link_success = False
                print(f"Download failed for {img_url}: {e}")
                return {
                    'Page': page, 'Index': idx if idx is not None else '?', 'Screenshot Link': link, 'Image Link': img_url,'Img Link Query': image_url_query,
                    'Content Disposition': '', 'Content Type': '', 'File Name': '', 'App ID': '',
                    'Missing Content Disposition': missing_cd, 'Same Content Disposition': False, 'Same File Name': False, 'File Written': False, 'Image Link Success': False
                }
    return link
# Check for duplicate Content-Disposition
def check_same_cd(existing_cds, cd_header):
    for existing in existing_cds:
        if existing == cd_header:
            return True
    return False

# Check for duplicate filename per app
def check_same_fname_per_app(existing_fnames, app_id, fname):
    key = f"{app_id}:{fname}"
    if key in existing_fnames:
        return True
    existing_fnames[key] = True
    return False

# Threaded download images with tracking
def threaded_download_imgs(img_links, page, save_dir, stop_flag, debug, max_retries=5, retry_history=None, 
                         processes=8, tracking=[], existing_cds=None, existing_fnames=None):
    if retry_history is None:
        retry_history = {}
    if existing_cds is None:
        existing_cds = []
    if existing_fnames is None:
        existing_fnames = {}

    error_links = []
    success_count = 0
    lock = threading.Lock()

    def wrapped_download(link_tuple):
        index, link, img_url, image_url_query = link_tuple

        result = download_img(page, link, img_url, image_url_query, save_dir, stop_flag, debug, index)
        
        if isinstance(result, dict):
            #check for duplicates and update tracking with lock
            with lock:
                result['Same Content Disposition'] = check_same_cd(existing_cds, result['Content Disposition'])
                if result['Same Content Disposition']:
                    existing_cds.append(result['Content Disposition'])
                result['Same File Name'] = check_same_fname_per_app(existing_fnames, result['App ID'], result['File Name'])

                if not result['Same File Name']:
                    existing_fnames[f"{result['App ID']}:{result['File Name']}"] = True
                tracking.append(result)
            return index, link, img_url, image_url_query, None
        else:
            return index, link, img_url, image_url_query, result

    with ThreadPoolExecutor(max_workers=processes) as executor:
        future_to_url = {
            executor.submit(wrapped_download, link): link
            for link in img_links
        }

        for future in as_completed(future_to_url):
            index, link, img_url, image_url_query, result = future.result()
            
            if result is None:
                with lock:
                    success_count += 1
                    print(f"[Page {page}] Downloaded {index+1} of {len(img_links)} screenshots...")
            else:
                retry_history[img_url] = retry_history.get(img_url, 0) + 1
                if retry_history[img_url] < max_retries:
                    error_links.append((index, link, img_url, image_url_query))
                else:
                    print(f"[image {index+1}]\n{img_url}\n[Full url]\n{link}")

    return error_links, retry_history, success_count

#Extract timestamp from filenames with new/old formats
def extract_datetime_from_filename(fname):
    # 1.new format(after 2016): 20250403215812_1.jpg
    match = re.match(r'^(\d{14})_\d+', fname)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y%m%d%H%M%S")
        except:
            pass
    # 3.new format(after 2016): 20250403215812.jpg
    match = re.match(r'^(\d{14})', fname)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y%m%d%H%M%S")
        except:
            pass
    # 2.old format(before 2016): 2012-01-24_00001.jpg
    match = re.match(r'^(\d{4}-\d{2}-\d{2})_', fname)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y-%m-%d")
        except:
            pass
    return None


def set_creation_time_linux(filepath, timestamp):

    if isinstance(timestamp, datetime):
        timestamp = timestamp.timestamp()
    try:
        os.utime(filepath, (timestamp, timestamp))  # (atime, mtime)
    except Exception as e:
        print(f"Failed to set modification time for {filepath}: {e}")

# # Set creation time for one file
# def set_creation_time(filepath, timestamp):
#     wintime = pywintypes.Time(timestamp)
#     handle = win32file.CreateFile(
#         filepath,
#         win32con.GENERIC_WRITE,
#         win32con.FILE_SHARE_READ | win32con.FILE_SHARE_WRITE | win32con.FILE_SHARE_DELETE,
#         None,
#         win32con.OPEN_EXISTING,
#         win32con.FILE_ATTRIBUTE_NORMAL,
#         None
#     )
#     win32file.SetFileTime(handle, wintime, wintime, wintime)
#     handle.close()

# Set creation times for all downloaded images
# Set creation times for all downloaded images
def set_all_creation_times(download_dir):
    artwork_dir = False
    error_dir = False
    for root, _, files in os.walk(download_dir):
        if "Artwork" in root.split(os.sep):
                if not artwork_dir:
                    print("Skipping Artwork directory for creation time update.")
                artwork_dir = True
                continue
         
        for fname in files:
            if fname.lower().endswith(".jpg"):
                dt = extract_datetime_from_filename(fname)
                if dt:
                    filepath = os.path.join(root, fname)
                    try:
                        set_creation_time_linux(filepath, dt)
                    except Exception as e:
                        print(f"Failed to set creation time for {fname}: {e}")
                else:
                    print(f"Skipped (no timestamp found): {fname}")

# Generate thumbnails
def generate_thumbnails(screenshots_dir, thumbnails_dir, size=(200, 125)):
    os.makedirs(thumbnails_dir, exist_ok=True)
    for fname in os.listdir(screenshots_dir):
        if not fname.lower().endswith(('.jpg', '.jpeg', '.png')):
            continue
        src_path = os.path.join(screenshots_dir, fname)
        dst_path = os.path.join(thumbnails_dir, fname)
        try:
            with Image.open(src_path) as img:
                img.thumbnail(size)
                img.save(dst_path, "JPEG", quality=85)
        except Exception as e:
            print(f"Failed to create thumbnail for {fname}: {e}")

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.abspath(relative_path)

# GUI App
class SteamDownloaderApp:
    def __init__(self, root):
        self.stop_flag = False
        self.root = root
        root.title("Steam Screenshot Downloader V2.9")

        # Menu bar
        menubar = tk.Menu(root)
        manual_menu = tk.Menu(menubar, tearoff=0)
        manual_menu.add_command(label="Manual Download", command=self.manual_download_window)
        menubar.add_cascade(label="Manual", menu=manual_menu)

        #help bar
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Common Questions", command=self.show_common_questions)
        help_menu.add_command(label="About", command=self.show_about_info)
        menubar.add_cascade(label="Help", menu=help_menu)


        root.config(menu=menubar)

        #root.iconbitmap(resource_path("Steam&Cookies.ico"))
        root.geometry("900x600")
        root.resizable(True, True)

        style = ttk.Style()
        style.configure("TLabel", font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10))

        # Variables
        self.steam_id = tk.StringVar()
        self.download_dir = tk.StringVar()
        self.chrome_dir = tk.StringVar()
        self.start_page = tk.StringVar()
        self.end_page = tk.StringVar()
        self.appid = tk.StringVar()
        self.processes = tk.StringVar()
        self.screenshot = tk.BooleanVar(value=True) 
        self.favorite = tk.BooleanVar(value=False)
        self.custom_url = tk.BooleanVar(value=True)
        self.debug = tk.BooleanVar(value=False)

        # Set default download directory
        default_download_dir = os.path.join(os.getcwd(), "Download_Screenshot")
        self.download_dir.set(default_download_dir)



        default_chrome_path = detect_firefox_default_profile_folder()
        if os.path.exists(default_chrome_path):
            self.chrome_dir.set(default_chrome_path)
        else:
            self.chrome_dir.set("")

        # # Set default Chrome user data path
        # user_profile = os.environ.get("USERPROFILE", "")
        # #default_chrome_path = os.path.join(user_profile, "AppData", "Local", "Google", "Chrome for Testing", "User Data")
        # #default_chrome_path = os.path.join(user_profile, ".config", "chromium")
        # default_chrome_path = "/home/jinx/.config/google-chrome-for-testing"
        # self.chrome_dir.set(default_chrome_path)
        # if os.path.exists(default_chrome_path):
        #     self.chrome_dir.set(default_chrome_path)

        self.processes.set(8)

        self.make_widgets()
        sys.stdout = TextRedirector(self.text_log)

    def make_widgets(self):
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill="both", expand=True)

        #Set up five columns: 50px fixed + left spacer + content + right spacer + 50px fixed
        frame.grid_columnconfigure(0, minsize=50)
        frame.grid_columnconfigure(1, minsize=15,weight=1)
        frame.grid_columnconfigure(2, weight=2)
        frame.grid_columnconfigure(3, weight=1)
        frame.grid_columnconfigure(4, minsize=50) 
        frame.rowconfigure(9, weight=1)

        # Steam ID
        ttk.Label(frame, text="Steam ID:").grid(row=0, column=1, sticky="e", pady=3, padx=(0, 5))
        ttk.Entry(frame, textvariable=self.steam_id).grid(row=0, column=2, sticky="ew", pady=3)
        ttk.Checkbutton(frame, text="Custom URL", variable=self.custom_url).grid(row=0, column=3, sticky="w", padx=(5, 0))
        # Save Location
        ttk.Label(frame, text="Save Location:").grid(row=1, column=1, sticky="e", pady=3, padx=(0, 5))
        ttk.Entry(frame, textvariable=self.download_dir).grid(row=1, column=2, sticky="ew", pady=3)
        ttk.Button(frame, text="Browse", command=self.select_download_dir, width=10).grid(row=1, column=3, pady=3)

        # Chrome Dir
        ttk.Label(frame, text="Firefox User Data:").grid(row=2, column=1, sticky="e", pady=3, padx=(0, 5))
        ttk.Entry(frame, textvariable=self.chrome_dir).grid(row=2, column=2, sticky="ew", pady=3)
        ttk.Button(frame, text="Browse", command=self.select_chrome_dir, width=10).grid(row=2, column=3, pady=3)

        # App ID + Max Processes
        ttk.Label(frame, text="App ID (optional):").grid(row=3, column=1, sticky="e", pady=3, padx=(0, 5))
        idproc_frame = ttk.Frame(frame)
        idproc_frame.grid(row=3, column=2, sticky="w", pady=3, columnspan=2)
        ttk.Entry(idproc_frame, textvariable=self.appid, width=20).pack(side="left")
        ttk.Label(idproc_frame, text="(leave blank for all)", foreground="gray").pack(side="left", padx=(5, 10))
        ttk.Label(idproc_frame, text="Threads:").pack(side="left")
        ttk.Entry(idproc_frame, textvariable=self.processes, width=10).pack(side="left", padx=(2, 0))
    
        # Start Page
        ttk.Label(frame, text="Start Page:").grid(row=4, column=1, sticky="e", pady=3, padx=(0, 5))
        ttk.Entry(frame, textvariable=self.start_page, width=10).grid(row=4, column=2, sticky="w", pady=3)

        # End Page
        ttk.Label(frame, text="End Page:").grid(row=5, column=1, sticky="e", pady=3, padx=(0, 5))
        ttk.Entry(frame, textvariable=self.end_page, width=10).grid(row=5, column=2, sticky="w", pady=3)

        # Screenshot or Artwork
        type_frame = ttk.Frame(frame)
        type_frame.grid(row=6, column=2, sticky="w", pady=3)
        ttk.Radiobutton(type_frame, text="Screenshot", variable=self.screenshot, value=True).pack(side="left", padx=0)
        ttk.Radiobutton(type_frame, text="Artwork", variable=self.screenshot, value=False).pack(side="left", padx=10)
        ttk.Frame(type_frame, width=150).pack(side="left")
        ttk.Radiobutton(type_frame, text="My Files", variable=self.favorite, value=False).pack(side="left", padx=0)
        ttk.Radiobutton(type_frame, text="My Favorite", variable=self.favorite, value=True).pack(side="left", padx=10)

        ttk.Checkbutton(frame, text="Debug Mode", variable=self.debug).grid(row=3, column=2, sticky="w", padx=(460, 0))

        # Browse button
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=8, column=2, pady=10)
        ttk.Button(button_frame, text="Start Download", width=15, command=self.start_thread).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Stop Download", width=15, command=self.stop_download).pack(side="left", padx=10)

        # Log output box
        self.text_log = tk.Text(frame, height=12, wrap="word", font=("Consolas", 9),
                                borderwidth=1, relief="solid", highlightthickness=0)
        self.text_log.grid(row=9, column=2, sticky="nsew", pady=10)
        frame.rowconfigure(9, weight=1)
        self.text_log.config(state="normal")

    def select_download_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.download_dir.set(path)

    def select_chrome_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.chrome_dir.set(path)

    def start_thread(self):
        t = threading.Thread(target=self.run_downloader)
        t.start()

    def stop_download(self):
        self.stop_flag = True
        print("⛔Stop after completing the remaining tasks...")
        time.sleep(3)
        self.stop_flag = False
    def show_common_questions(self):
        #Common Questions window
        win = tk.Toplevel(self.root)
        win.title("Common Questions")
        win.geometry("650x450")
        #win.iconbitmap(resource_path("Steam&Cookies.ico"))

        text = tk.Text(win, wrap="word", font=("Segoe UI", 10), bg="#f8f8f8")
        text.insert("1.0",
            "📘 Common Questions:\n\n"
            "1. How to Use?\n"
            "   →1. After clicking Download, log in to Steam in the Chromium window that appears the first time.\n"
            "     2. Once the download is complete, you can check the Tracking Report.\n"
            "     3. Any failed links can be copied directly into Manual Download.\n\n"

            "2. Why can't it download (the browser flashes briefly)?\n"
            "   → The tool relies on your local Chromium installation. Try placing the program on the C: drive(e.g., Desktop).\n\n"

            "3. Why log in to Steam in Chromium?\n"
            "   → Downloading your private screenshots requires Steam cookies, and using the browser helps avoid manually entering long strings.\n\n"

            "4. How to use manual download?\n"
            "   → Click 'Manual' → 'Manual Download', paste screenshot links, then press 'Download'.\n\n"

            "5. Where are downloaded screenshots saved?\n"
            "   → Default path: ./Download_Screenshot folder.\n\n"
            "For a detailed user guide, please visit the GitHub page.\n"
            "https://github.com/notcookies/Steam-Screenshot-downloader.\n\n"

        )
        text.config(state="disabled")  # 只读
        text.pack(expand=True, fill="both", padx=10, pady=10)


    def show_about_info(self):
        # About window
        win = tk.Toplevel(self.root)
        win.title("About Steam Screenshot Downloader")
        win.geometry("400x250")
        #win.iconbitmap(resource_path("Steam&Cookies.ico"))

        tk.Label(
            win,
            text="Steam Screenshot Downloader\nVersion 2.9",
            font=("Segoe UI", 12, "bold")
        ).pack(pady=10)

        tk.Label(
            win,
            text=
            "A lightweight GUI tool for downloading\n"
            "your uploaded or favorited\n"
            "Public/Private Steam Screenshots/Artworks\n"
            "or content shared by users.",
            font=("Segoe UI", 10),
            justify="center"
        ).pack(pady=10)

        link = tk.Label(
            win,
            text="GitHub Repository",
            font=("Segoe UI", 10, "underline"),
            fg="blue",
            cursor="hand2"
        )
        link.pack(pady=10)

        # open github link
        def open_github(event):
            import webbrowser
            webbrowser.open("https://github.com/notcookies/Steam-Screenshot-downloader")

        link.bind("<Button-1>", open_github)

        tk.Label(win, text="© 2025 Notcookies. All rights reserved.", font=("Segoe UI", 8)).pack(side="bottom", pady=10)

    def tracking_report(self, tracking, download_dir):
            
        tracking_df = pd.DataFrame(tracking)

        #create excel report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(download_dir, f"Tracking_Report_{timestamp}.xlsx")

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # mian tracking data              
            # app id as int
            # sort by page and index
            tracking_df = tracking_df.sort_values(['Page', 'Index']).reset_index(drop=True)
            tracking_df['Index'] = tracking_df.groupby('Page').cumcount() + 1

            tracking_df.to_excel(writer, sheet_name='Tracking_Data', index=False)
            
            #statistics summary
            summary_data = {
                'Category': [
                    'Total Records:',
                    'Successful img_link Parsing:', 
                    'Missing Content-Disposition:',
                    'Duplicate Content-Disposition:',
                    'File Write Success:',
                    'Duplicate Filename Count:'
                ],
                'Count': [
                    len(tracking_df),
                    tracking_df['Image Link Success'].sum(),
                    tracking_df['Missing Content Disposition'].sum(),
                    tracking_df['Same Content Disposition'].sum(),
                    tracking_df['File Written'].sum(),
                    tracking_df['Same File Name'].sum()
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # App stats
            app_stats = tracking_df[tracking_df['File Written'] == True].groupby('App ID').agg({
                'File Name': 'count',
                'Missing Content Disposition': 'sum',
                'Same File Name': 'sum'
            }).rename(columns={'File Name': 'Screenshots Cont',})
            app_stats.to_excel(writer, sheet_name='App_Stats')
            
            #Error report
            issues_df = tracking_df[
                (tracking_df['Image Link Success'] == False) | 
                (tracking_df['Missing Content Disposition'] == True) | 
                (tracking_df['Same Content Disposition'] == True) | 
                (tracking_df['Same File Name'] == True) | 
                (tracking_df['File Written'] == False)
            ][['Page', 'Index', 'App ID', 'File Name', 'Image Link', 
            'Image Link Success', 'Missing Content Disposition', 'Same Content Disposition', 'Same File Name', 'File Written']]
            issues_df.to_excel(writer, sheet_name='Issues', index=False)
            
            #report duplicate img_links
            duplicates = tracking_df[tracking_df.duplicated('Image Link', keep=False)]
            if not duplicates.empty:
                duplicates[['Page', 'Index', 'App ID', 'Image Link']].to_excel(
                    writer, sheet_name='Duplicate_img_links', index=False)
            
            # check actual files on disk for those marked as written
            missing_files = []
            for _, row in tracking_df.iterrows():
                if row['File Written'] == True:
                    path = os.path.join(download_dir, row['App ID'], "screenshots", row['File Name'])
                    actual_exists = os.path.exists(path) and os.path.getsize(path) > 0
                    if not actual_exists:
                        missing_files.append({
                            'Page': row['Page'], 
                            'Index': row['Index'], 
                            'App ID': row['App ID'], 
                            'File Name': row['File Name'],
                            'Expected Path': path,
                            'Image Link': row['Image Link']
                        })
            
            if missing_files:
                missing_df = pd.DataFrame(missing_files)
                missing_df.to_excel(writer, sheet_name='Missing_Files', index=False)
            
        workbook = writer.book

        def auto_fit_columns(worksheet):
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = max(width, 12)

        #apply auto fit to all sheets
        for sheet_name in workbook.sheetnames:
            auto_fit_columns(workbook[sheet_name])

        workbook.save(excel_file)
                
        #all tracking report done
        print(f"\n✅all tracking report done: {excel_file}")

        # format Excel
        workbook = writer.book
        worksheet = workbook['Tracking_Data']
        from openpyxl.styles import Font, PatternFill, Alignment

        # format header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(excel_file)

        #Total records, successful img_links, file writes, issues, missing files, app stats
        print(f"Total records: {len(tracking_df)}")
        print(f"Successful image links: {tracking_df['Image Link Success'].sum()}")
        print(f"File writes: {tracking_df['File Written'].sum()}")
        print(f"Error image info: {len(issues_df)}")
        if missing_files:
            print(f"Missing files: {len(missing_files)}")
        print(f"App count: {len(app_stats)}")

    def run_downloader(self):
        steam_id = self.steam_id.get().strip()
        download_dir = os.path.normpath(self.download_dir.get().strip())
        chrome_dir = os.path.normpath(self.chrome_dir.get().strip())
        appid_input = self.appid.get().strip()
        appid = appid_input if appid_input else None
        processes_input = self.processes.get().strip()
        processes = processes_input if processes_input else 8
        processes = int(processes)

        tracking = []

        try:
            start_page = int(self.start_page.get())
            end_page = int(self.end_page.get())
        except:
            print("Start and end pages must be integers.")
            return

        if not (steam_id and download_dir and chrome_dir):
            print("All fields are required.")
            return

        if not os.path.exists(download_dir):
            os.makedirs(download_dir, exist_ok=True)

        driver = init_chrome(chrome_dir)
        time.sleep(random.uniform(1, 2))
        
        cookies = None

        for page in range(start_page, end_page + 1):
            if self.stop_flag:
                print("⛔ Download stopped by user.")
                break

            #get screenshot links with original index
            indexed_links = get_screenshot_links(driver, steam_id, page, appid=appid, 
                                            screenshot=self.screenshot.get(), 
                                            favorite=self.favorite.get(), 
                                            custom_url=self.custom_url.get())
            if not indexed_links:
                continue

            # try get cookies
            if cookies is None:
                print("Extracting Steam cookies from Firefox session...")
                cookies = extract_steam_cookies_from_driver(driver, retries=3, delay=2)
                if not all(k in cookies for k in ["steamLoginSecure", "sessionid"]):
                    print("❌ Missing required cookies from Firefox session.")
                    print("Please ensure you\'re logged into Steam in Chrome.")
                    print("Chrome will Close in 5 minutes.....")
                    time.sleep(300)
                    driver.quit()
                    return
                
            #from links get img_links with original index
            img_links = fetch_img_urls_concurrently_requests(
                indexed_links, cookies, page, processes, max_retries=3
            )
            
            if not img_links:
                continue
                
            successful_downloads = 0
            retry_links = img_links
            retry_history = {}
            max_retries = 4
            attempt = 0

            #link_to_idx = {link: idx for idx, link in enumerate(indexed_links)}
            #link_to_idx = {link: index for index, link in indexed_links}

            while retry_links and attempt < max_retries:
                attempt += 1
                if attempt == 2:
                    print(f"\nFind {len(retry_links)} Error screenshot links.")
                if attempt > 1:
                    print(f"🔁 Retrying attempt {attempt-1}.")
                if attempt == max_retries:
                    print('Failed to download after max retries:')
                    
                # original_indices
                retry_links, retry_history, new_success = threaded_download_imgs(
                    retry_links, page, download_dir, self.stop_flag, self.debug.get(), 
                    max_retries, retry_history, processes=processes, tracking=tracking
                )
                successful_downloads += new_success

            print(f"\n[Page {page}] ✅ Page download completed. Total downloaded: {successful_downloads}.\n")
        
        driver.quit()
        print("✅ All downloads completed.")

        # add tracking info and check duplicates
        # generate excel report
        if tracking:
            self.tracking_report(tracking, download_dir)
        else:
            print("No tracking data generated.")


        set_all_creation_times(download_dir)
        print("✅ Creation time update completed.")

        for appid in os.listdir(download_dir):
            if appid == "Artwork":
                print("Skipping thumbnails for Artwork.")
            else:
                app_screenshot_dir = os.path.join(download_dir, appid, "screenshots")
                app_thumbnail_dir = os.path.join(app_screenshot_dir, "thumbnails")
                if os.path.isdir(app_screenshot_dir):
                    generate_thumbnails(app_screenshot_dir, app_thumbnail_dir)
        print("✅ Thumbnails generation completed.")
        print("✅ All Done.")


    def manual_download_window(self):
        win = tk.Toplevel(self.root)
        win.title("Manual Screenshot Link Download")
        win.geometry("600x600")
        #win.iconbitmap(resource_path("Steam&Cookies.ico"))

        win.rowconfigure(1, weight=1)
        win.columnconfigure(0, weight=1)

        # Placeholder content
        placeholder = (
            "Enter Steam Screenshot Links, such as:\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=1346814516\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=1376701612\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=1376701717\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=2837849631\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=2837849646\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=2837849692\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=2837849756\n"
            "https://steamcommunity.com/sharedfiles/filedetails/?id=2837849767\n"
            "......\n"
        )

        # Text box with placeholder
        text_box = tk.Text(win, wrap="word", font=("Consolas", 10), height=10, fg="gray")
        text_box.insert("1.0", placeholder)
        text_box.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

        # Focus in to clear placeholder
        def on_focus_in(event):
            if text_box.get("1.0", "end-1c") == placeholder:
                text_box.delete("1.0", "end")
                text_box.config(fg="black")

        # Focus out to restore placeholder if empty
        def on_focus_out(event):
            if not text_box.get("1.0", "end-1c").strip():
                text_box.insert("1.0", placeholder)
                text_box.config(fg="gray")

        text_box.bind("<FocusIn>", on_focus_in)
        text_box.bind("<FocusOut>", on_focus_out)

        #  start a new thread to download for avoid blocking GUI
        def start_manual_download():
            links_text = text_box.get("1.0", tk.END)
            # avoid empty or placeholder input
            if links_text.strip() == placeholder.strip() or not links_text.strip():
                print("⚠️ Please enter at least one valid screenshot link.")
                return
            threading.Thread(
                target=self.manual_download,
                args=(links_text,),
                daemon=True
            ).start()

        ttk.Button(win, text="Download", command=start_manual_download).grid(row=2, column=0, pady=10)


    def manual_download(self, links_text):
        # Clean and split links from text box
        links = [l.strip() for l in links_text.strip().splitlines() if l.strip()]
        if not links:
            print("❌ Please enter Steam screenshot link.")
            return
        
        # Validate link format
        valid_links = [l for l in links if l.startswith("https://steamcommunity.com/sharedfiles/filedetails/")]
        if not valid_links:
            print("❌ No Steam screenshot links found.eg., https://steamcommunity.com/sharedfiles/filedetails/?id=......")
            return
        
        tracking = []

        page = 0
        save_dir = os.path.join(self.download_dir.get(), "Manual")
        os.makedirs(save_dir, exist_ok=True)

        debug = self.debug.get()
        stop_flag = self.stop_flag
        processes = int(self.processes.get() or 8)

        print(f"Manual mode: downloading {len(valid_links)} screenshots...")

        chrome_dir = os.path.normpath(self.chrome_dir.get().strip())

        driver = init_chrome(chrome_dir)
        print("Fetch cookies from Steam Community....")
        driver.get("https://steamcommunity.com/")
        
        cookies = extract_steam_cookies_from_driver(driver, retries=3, delay=2)
        if not all(k in cookies for k in ["steamLoginSecure", "sessionid"]):
            print("❌ Missing required cookies from Chrome session.")
            print("Please ensure you\'re logged into Steam in Chrome.")
            print("Chrome will Close in 5 minutes.....")
            time.sleep(300)
            driver.quit()
            return
        
        indexed_links = list(enumerate(valid_links))
        
        img_links = fetch_img_urls_concurrently_requests(
            indexed_links, cookies, page, processes=processes, max_retries=3
        )

        if not img_links:
            print(f"❌ Failed to fetch image URLs for {valid_links}. Skipping...")
            driver.quit()
            return
        
        
        threaded_download_imgs(
            img_links, page, save_dir, stop_flag, debug, max_retries=5, 
            retry_history=None, processes=processes, 
            existing_cds=None, existing_fnames=None,tracking=tracking
        )

        driver.quit()
        print("✅ All manual downloads completed.\n")
        if tracking:
            self.tracking_report(tracking, save_dir)
        else:
            print("No tracking data generated.")

        set_all_creation_times(save_dir)
        print("✅ Creation time update completed.")

        for appid in os.listdir(save_dir):
            if appid == "Artwork":
                print("Skipping thumbnails for Artwork.")
            else:
                app_screenshot_dir = os.path.join(save_dir, appid, "screenshots")
                app_thumbnail_dir = os.path.join(app_screenshot_dir, "thumbnails")
                if os.path.isdir(app_screenshot_dir):
                    generate_thumbnails(app_screenshot_dir, app_thumbnail_dir)
        print("✅ Thumbnails generation completed.")
        print("✅ All Done.")



# Run GUI
if __name__ == "__main__":
    root = tk.Tk()
    app = SteamDownloaderApp(root)
    root.mainloop()